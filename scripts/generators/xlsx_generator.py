"""Generate synthetic Excel (.xlsx) test files using openpyxl."""

from pathlib import Path

from generators.base import BaseGenerator

try:
    import openpyxl
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    _OXL_OK = True
except ImportError:
    _OXL_OK = False


def _require_openpyxl() -> None:
    if not _OXL_OK:
        raise ImportError("openpyxl is required: uv add openpyxl")


def _header_fill(hex_color: str = "1F4E79") -> "PatternFill":
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _thin_border() -> "Border":
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


class XlsxGenerator(BaseGenerator):
    category = "xlsx"

    def generate(self, dry_run: bool = False) -> list[dict]:
        _require_openpyxl()
        out = self.output_dir()
        if not dry_run:
            out.mkdir(parents=True, exist_ok=True)

        entries = []
        entries.append(self._simple_grid(out, dry_run))
        entries.append(self._multi_sheet(out, dry_run))
        entries.append(self._merged_cells(out, dry_run))
        entries.append(self._chart_embedded(out, dry_run))
        entries.append(self._wide_sheet(out, dry_run))
        entries.append(self._rich_formatting(out, dry_run))
        return entries

    # ------------------------------------------------------------------
    def _simple_grid(self, out: Path, dry_run: bool) -> dict:
        filename = "syn_simple_grid.xlsx"
        if dry_run:
            print(f"  [dry-run] would create {self.category}/{filename}")
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sales Data"
            headers = ["Quarter", "Region", "Product", "Units Sold", "Revenue ($)", "Margin (%)"]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = _header_fill()

            rows = [
                ("Q1", "North", "Widget A", 1200, 48000, 32.5),
                ("Q1", "South", "Widget B", 890, 31150, 28.1),
                ("Q2", "North", "Widget A", 1350, 54000, 33.0),
                ("Q2", "East", "Widget C", 670, 23450, 25.8),
                ("Q3", "West", "Widget A", 1100, 44000, 31.5),
                ("Q3", "South", "Widget B", 950, 33250, 29.2),
                ("Q4", "North", "Widget C", 800, 28000, 27.0),
                ("Q4", "East", "Widget A", 1420, 56800, 34.1),
                ("Q4", "West", "Widget B", 760, 26600, 26.8),
                ("Q4", "South", "Widget C", 880, 30800, 27.5),
            ]
            for row in rows:
                ws.append(row)
            wb.save(out / filename)
        return self._entry(filename, "easy",
                           ["xlsx", "single-sheet", "simple-grid", "tabular-data"],
                           False, "Simple 10-row data grid, single sheet, basic header formatting.")

    def _multi_sheet(self, out: Path, dry_run: bool) -> dict:
        filename = "syn_multi_sheet.xlsx"
        if dry_run:
            print(f"  [dry-run] would create {self.category}/{filename}")
        else:
            wb = openpyxl.Workbook()

            # Sheet 1: Summary
            ws1 = wb.active
            ws1.title = "Summary"
            ws1.append(["Metric", "Q1", "Q2", "Q3", "Q4", "Full Year"])
            ws1.append(["Revenue ($K)", 820, 940, 1050, 1180, "=SUM(B2:E2)"])
            ws1.append(["Expenses ($K)", 610, 680, 720, 790, "=SUM(B3:E3)"])
            ws1.append(["Net Profit ($K)", "=B2-B3", "=C2-C3", "=D2-D3", "=E2-E3", "=SUM(B4:E4)"])
            ws1.append(["Margin (%)", "=B4/B2*100", "=C4/C2*100", "=D4/D2*100", "=E4/E2*100", "=F4/F2*100"])
            for cell in ws1[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = _header_fill()

            # Sheet 2: Monthly Detail
            ws2 = wb.create_sheet("Monthly Detail")
            ws2.append(["Month", "Revenue", "COGS", "Gross Profit", "Operating Expenses", "Net Income"])
            months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                      "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
            base_rev = 250
            for i, m in enumerate(months):
                rev = base_rev + i * 15 + (i % 3) * 10
                cogs = int(rev * 0.45)
                gross = rev - cogs
                opex = int(rev * 0.30)
                net = gross - opex
                ws2.append([m, rev * 1000, cogs * 1000, gross * 1000, opex * 1000, net * 1000])
            for cell in ws2[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = _header_fill("2E75B6")

            # Sheet 3: Notes
            ws3 = wb.create_sheet("Notes & Assumptions")
            ws3["A1"] = "Financial Model Assumptions"
            ws3["A1"].font = Font(bold=True, size=14)
            assumptions = [
                ("Revenue growth rate:", "6% month-over-month"),
                ("COGS as % of revenue:", "45%"),
                ("Operating expense ratio:", "30%"),
                ("Tax rate:", "Not applied in this model"),
                ("Currency:", "USD thousands"),
                ("Fiscal year:", "January – December 2024"),
            ]
            for i, (label, val) in enumerate(assumptions, start=3):
                ws3[f"A{i}"] = label
                ws3[f"B{i}"] = val
                ws3[f"A{i}"].font = Font(bold=True)

            wb.save(out / filename)
        return self._entry(filename, "medium",
                           ["xlsx", "multi-sheet", "formulas", "cross-sheet", "financial-model"],
                           False, "Three-sheet workbook: Summary with formulas, Monthly Detail, and Notes.")

    def _merged_cells(self, out: Path, dry_run: bool) -> dict:
        filename = "syn_merged_cells.xlsx"
        if dry_run:
            print(f"  [dry-run] would create {self.category}/{filename}")
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Department Budget"

            # Title row spanning all columns
            ws.merge_cells("A1:F1")
            ws["A1"] = "Annual Department Budget — FY2024"
            ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
            ws["A1"].fill = _header_fill("1F4E79")
            ws["A1"].alignment = Alignment(horizontal="center")

            # Section header: Q1/Q2 spanning
            ws.merge_cells("B2:C2")
            ws["B2"] = "H1 2024"
            ws["B2"].alignment = Alignment(horizontal="center")
            ws.merge_cells("D2:E2")
            ws["D2"] = "H2 2024"
            ws["D2"].alignment = Alignment(horizontal="center")
            for cell in ("B2", "D2"):
                ws[cell].font = Font(bold=True, color="FFFFFF")
                ws[cell].fill = _header_fill("2E75B6")

            ws["A3"] = "Department"
            ws["B3"] = "Q1 Budget"
            ws["C3"] = "Q2 Budget"
            ws["D3"] = "Q3 Budget"
            ws["E3"] = "Q4 Budget"
            ws["F3"] = "Annual Total"
            for col in "ABCDEF":
                ws[f"{col}3"].font = Font(bold=True)
                ws[f"{col}3"].fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

            depts = [
                ("Engineering", 450000, 480000, 510000, 530000),
                ("Marketing", 180000, 220000, 200000, 250000),
                ("Sales", 320000, 340000, 360000, 410000),
                ("HR & Admin", 120000, 125000, 122000, 130000),
                ("Finance", 95000, 98000, 100000, 105000),
                ("Operations", 280000, 290000, 310000, 320000),
            ]
            for i, (dept, q1, q2, q3, q4) in enumerate(depts, start=4):
                ws[f"A{i}"] = dept
                ws[f"B{i}"] = q1
                ws[f"C{i}"] = q2
                ws[f"D{i}"] = q3
                ws[f"E{i}"] = q4
                ws[f"F{i}"] = f"=SUM(B{i}:E{i})"

            # Total row with merged label
            last = len(depts) + 4
            ws[f"A{last}"] = "TOTAL"
            ws[f"A{last}"].font = Font(bold=True)
            for col in "BCDEF":
                ws[f"{col}{last}"] = f"=SUM({col}4:{col}{last - 1})"
                ws[f"{col}{last}"].font = Font(bold=True)

            wb.save(out / filename)
        return self._entry(filename, "medium",
                           ["xlsx", "merged-cells", "department-budget", "section-headers"],
                           False, "Budget spreadsheet with merged title row and merged column-group headers.")

    def _chart_embedded(self, out: Path, dry_run: bool) -> dict:
        filename = "syn_chart_embedded.xlsx"
        if dry_run:
            print(f"  [dry-run] would create {self.category}/{filename}")
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sales Performance"

            ws.append(["Month", "Product A", "Product B", "Product C"])
            data = [
                ("Jan", 45, 32, 28),
                ("Feb", 52, 38, 31),
                ("Mar", 61, 42, 35),
                ("Apr", 58, 45, 40),
                ("May", 70, 50, 38),
                ("Jun", 75, 55, 44),
                ("Jul", 68, 48, 41),
                ("Aug", 82, 60, 49),
                ("Sep", 79, 57, 52),
                ("Oct", 88, 63, 55),
                ("Nov", 95, 70, 60),
                ("Dec", 102, 78, 67),
            ]
            for row in data:
                ws.append(row)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = _header_fill()

            # Create bar chart
            chart = BarChart()
            chart.type = "col"
            chart.grouping = "clustered"
            chart.title = "Monthly Sales by Product"
            chart.y_axis.title = "Units Sold (thousands)"
            chart.x_axis.title = "Month"
            chart.width = 18
            chart.height = 12

            data_ref = Reference(ws, min_col=2, max_col=4, min_row=1, max_row=13)
            cats = Reference(ws, min_col=1, min_row=2, max_row=13)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, "F2")

            wb.save(out / filename)
        return self._entry(filename, "medium",
                           ["xlsx", "embedded-chart", "bar-chart", "sales-data"],
                           False, "Excel with 12-month sales data and an embedded clustered bar chart.")

    def _wide_sheet(self, out: Path, dry_run: bool) -> dict:
        filename = "syn_wide_100cols.xlsx"
        if dry_run:
            print(f"  [dry-run] would create {self.category}/{filename}")
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sensor Readings"

            headers = ["timestamp"] + [f"sensor_{i:03d}" for i in range(1, 100)]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)

            import datetime
            base = datetime.datetime(2024, 1, 1, 0, 0)
            for r in range(500):
                ts = base + datetime.timedelta(minutes=r * 5)
                row = [ts.isoformat()] + [round((r + i * 0.1) % 100, 2) for i in range(99)]
                ws.append(row)

            # Auto-size first column
            ws.column_dimensions["A"].width = 22

            wb.save(out / filename)
        return self._entry(filename, "hard",
                           ["xlsx", "wide-format", "100-columns", "500-rows", "time-series", "sensor-data"],
                           False, "Very wide sheet: 100 columns (timestamp + 99 sensor metrics), 500 rows.")

    def _rich_formatting(self, out: Path, dry_run: bool) -> dict:
        filename = "syn_rich_formatting.xlsx"
        if dry_run:
            print(f"  [dry-run] would create {self.category}/{filename}")
        else:
            from openpyxl.styles import numbers
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Portfolio Summary"

            ws.append(["Asset", "Type", "Purchase Price", "Current Value", "Gain/Loss", "Return %", "Weight %"])
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = _header_fill("1F4E79")

            assets = [
                ("AAPL", "Equity", 145.20, 182.50),
                ("MSFT", "Equity", 280.00, 415.20),
                ("BND", "Bond ETF", 75.10, 72.40),
                ("GLD", "Commodity", 168.00, 210.30),
                ("AMZN", "Equity", 95.00, 178.25),
                ("TLT", "Bond ETF", 92.50, 84.10),
                ("VNQ", "REIT", 85.30, 79.60),
                ("QQQ", "Equity ETF", 310.00, 440.80),
            ]
            total_current = sum(a[3] for a in assets)

            for i, (name, atype, purchase, current) in enumerate(assets, start=2):
                gain = current - purchase
                ret_pct = gain / purchase * 100
                weight = current / total_current * 100

                ws[f"A{i}"] = name
                ws[f"B{i}"] = atype
                ws[f"C{i}"] = purchase
                ws[f"D{i}"] = current
                ws[f"E{i}"] = gain
                ws[f"F{i}"] = ret_pct / 100  # stored as decimal for % format
                ws[f"G{i}"] = weight / 100

                # Color gain/loss green or red
                gain_color = "00AA00" if gain >= 0 else "CC0000"
                ws[f"E{i}"].font = Font(color=gain_color, bold=True)
                ws[f"F{i}"].font = Font(color=gain_color)

                # Number formats
                ws[f"C{i}"].number_format = '"$"#,##0.00'
                ws[f"D{i}"].number_format = '"$"#,##0.00'
                ws[f"E{i}"].number_format = '"$"#,##0.00;[Red]"$"-#,##0.00'
                ws[f"F{i}"].number_format = '0.00%'
                ws[f"G{i}"].number_format = '0.00%'

                # Alternating row background
                if i % 2 == 0:
                    bg = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
                    for col in "ABCDEFG":
                        ws[f"{col}{i}"].fill = bg

            wb.save(out / filename)
        return self._entry(filename, "hard",
                           ["xlsx", "rich-formatting", "conditional-colors", "number-formats", "portfolio"],
                           False, "Portfolio sheet with multiple number formats, color-coded gain/loss, alternating row shading.")
